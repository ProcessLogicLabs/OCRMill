"""
Section 232 Steel/Aluminum Declaration Exporter for OCRMill

Reads processed invoice CSVs, enriches with material composition data from parts database,
expands rows by material type, and exports Section 232-compliant Excel files.

Based on TariffMill's Section 232 export implementation.
"""

import csv
import sqlite3
from pathlib import Path
from typing import Dict, List, Optional, Tuple
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


class Section232Exporter:
    """
    Section 232 Export Handler

    Process:
    1. Read processed CSV files
    2. Lookup material composition from parts database
    3. Expand rows by material type (steel, aluminum, copper, wood, auto, non_232)
    4. Assign declaration flags and calculate proportional values
    5. Export to Excel with proper formatting
    """

    # Material type configuration (content_type, ratio_column, dec_code, color)
    MATERIAL_CONFIGS = [
        ('steel', 'steel_pct', '08', '4a4a4a'),          # Dark gray
        ('aluminum', 'aluminum_pct', '07', '6495ED'),    # Cornflower blue
        ('copper', 'copper_pct', '11', 'B87333'),        # Copper
        ('wood', 'wood_pct', '10', '8B4513'),            # Saddle brown
        ('auto', 'auto_pct', '', '2F4F4F'),              # Dark slate gray
        ('non_232', 'non_steel_pct', '', 'FF0000'),      # Red
    ]

    # Export column headers
    EXPORT_COLUMNS = [
        'Product No', 'ValueUSD', 'HTSCode', 'MID', 'Qty1', 'Qty2',
        'DecTypeCd', 'CountryofMelt', 'CountryOfCast', 'PrimCountryOfSmelt',
        'DeclarationFlag', 'SteelRatio', 'AluminumRatio', 'CopperRatio',
        'WoodRatio', 'AutoRatio', 'NonSteelRatio', 'DualDeclaration', '232_Status', 'CustomerRef'
    ]

    def __init__(self, input_folder: Path, output_folder: Path, db_path: Path):
        """
        Initialize Section 232 Exporter.

        Args:
            input_folder: Folder containing processed CSV files
            output_folder: Folder for Section 232 exports
            db_path: Path to parts database
        """
        self.input_folder = Path(input_folder)
        self.output_folder = Path(output_folder)
        self.db_path = Path(db_path)
        self.output_folder.mkdir(parents=True, exist_ok=True)

    def process_all(self) -> int:
        """Process all CSV files in input folder and generate Section 232 exports."""
        csv_files = list(self.input_folder.glob("*.csv"))
        processed_count = 0

        for csv_file in csv_files:
            try:
                self.process_file(csv_file)
                processed_count += 1
            except Exception as e:
                print(f"Error processing {csv_file.name}: {e}")

        return processed_count

    def process_file(self, csv_path: Path):
        """Process a single CSV file and generate Section 232 export."""
        # Read CSV
        items = self._read_csv(csv_path)
        if not items:
            return

        # Enrich with material composition from database
        enriched_items = self._enrich_with_materials(items)

        # Expand rows by material type
        expanded_rows = self._expand_by_material(enriched_items)

        # Generate Excel export
        output_filename = f"232_{csv_path.stem}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        output_path = self.output_folder / output_filename
        self._export_to_excel(expanded_rows, output_path)

        print(f"Exported: {output_filename} ({len(expanded_rows)} rows)")

    def _read_csv(self, csv_path: Path) -> List[Dict]:
        """Read CSV file and return list of dictionaries."""
        items = []
        with open(csv_path, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            for row in reader:
                items.append(dict(row))
        return items

    def _enrich_with_materials(self, items: List[Dict]) -> List[Dict]:
        """Lookup material composition from parts database (TariffMill schema)."""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()

        for item in items:
            part_number = item.get('part_number', '') or item.get('Part Number', '')
            if part_number:
                cursor.execute("""
                    SELECT
                        steel_ratio, aluminum_ratio, non_steel_ratio, qty_unit
                    FROM parts_master
                    WHERE part_number = ?
                """, (part_number,))

                result = cursor.fetchone()
                if result:
                    # TariffMill uses ratio (0-100), map to expected column names
                    item['steel_pct'] = result[0] or 0.0
                    item['aluminum_pct'] = result[1] or 0.0
                    item['copper_pct'] = 0.0  # Not in TariffMill schema
                    item['wood_pct'] = 0.0  # Not in TariffMill schema
                    item['auto_pct'] = 0.0  # Not in TariffMill schema
                    item['non_steel_pct'] = result[2] or 0.0
                    item['qty_unit'] = result[3] or 'NO'
                else:
                    # Default to 0 if not found
                    item['steel_pct'] = 0.0
                    item['aluminum_pct'] = 0.0
                    item['copper_pct'] = 0.0
                    item['wood_pct'] = 0.0
                    item['auto_pct'] = 0.0
                    item['non_steel_pct'] = 0.0
                    item['qty_unit'] = 'NO'

        conn.close()
        return items

    def _expand_by_material(self, items: List[Dict]) -> List[Dict]:
        """
        Expand each item into multiple rows based on material composition.

        Logic from TariffMill:
        - For each material with ratio > 0, create a derivative row
        - Proportionally distribute value based on material percentage
        - Assign declaration flag based on material type
        """
        expanded = []

        for item in items:
            # Get original values
            original_value = float(item.get('total_price', 0) or item.get('Total Price', 0) or 0)
            net_weight = float(item.get('net_weight', 0) or item.get('Net Weight', 0) or 0)
            quantity = float(item.get('quantity', 0) or item.get('Quantity', 0) or 0)

            # Material percentages (0-100 scale)
            materials = {
                'steel_pct': float(item.get('steel_pct', 0) or 0),
                'aluminum_pct': float(item.get('aluminum_pct', 0) or 0),
                'copper_pct': float(item.get('copper_pct', 0) or 0),
                'wood_pct': float(item.get('wood_pct', 0) or 0),
                'auto_pct': float(item.get('auto_pct', 0) or 0),
                'non_steel_pct': float(item.get('non_steel_pct', 0) or 0),
            }

            # Check for dual declaration (both steel and aluminum > 0)
            has_steel = materials['steel_pct'] > 0
            has_aluminum = materials['aluminum_pct'] > 0
            dual_declaration = '07 & 08' if (has_steel and has_aluminum) else ''

            # Expand into derivative rows
            for content_type, ratio_key, dec_code, color in self.MATERIAL_CONFIGS:
                pct = materials[ratio_key]
                if pct > 0:
                    new_row = self._create_derivative_row(
                        item, content_type, pct, dec_code,
                        original_value, net_weight, quantity,
                        materials, dual_declaration
                    )
                    expanded.append(new_row)

        return expanded

    def _create_derivative_row(self, item: Dict, content_type: str, pct: float,
                                dec_code: str, original_value: float, net_weight: float,
                                quantity: float, materials: Dict, dual_declaration: str) -> Dict:
        """Create a single derivative row for a material type."""
        # Calculate proportional value
        proportional_value = original_value * pct / 100.0
        proportional_weight = net_weight * pct / 100.0

        # Determine declaration flag
        flag_map = {
            'steel': '232_Steel',
            'aluminum': '232_Aluminum',
            'copper': '232_Copper',
            'wood': '232_Wood',
            'auto': '232_Auto',
            'non_232': 'Non_232',
        }
        dec_flag = flag_map.get(content_type, '')

        # Determine Qty1 and Qty2 based on unit type
        qty_unit = item.get('qty_unit', 'NO')
        qty1, qty2 = self._calculate_quantities(qty_unit, quantity, proportional_weight, content_type)

        # Get country information
        country = item.get('country_origin', '') or item.get('Country Origin', '')

        # Build derivative row
        row = {
            'Product No': item.get('part_number', '') or item.get('Part Number', ''),
            'ValueUSD': f"{proportional_value:.2f}",
            'HTSCode': item.get('hts_code', '') or item.get('HTS Code', ''),
            'MID': item.get('mid', '') or item.get('MID', ''),
            'Qty1': qty1,
            'Qty2': qty2,
            'DecTypeCd': dec_code,
            'CountryofMelt': country if content_type in ['steel', 'aluminum', 'copper'] else '',
            'CountryOfCast': country if content_type in ['aluminum'] else '',
            'PrimCountryOfSmelt': country if content_type in ['aluminum', 'copper', 'wood'] else '',
            'DeclarationFlag': dec_flag,
            'SteelRatio': f"{materials['steel_pct']:.2f}%",
            'AluminumRatio': f"{materials['aluminum_pct']:.2f}%",
            'CopperRatio': f"{materials['copper_pct']:.2f}%",
            'WoodRatio': f"{materials['wood_pct']:.2f}%",
            'AutoRatio': f"{materials['auto_pct']:.2f}%",
            'NonSteelRatio': f"{materials['non_steel_pct']:.2f}%",
            'DualDeclaration': dual_declaration,
            '232_Status': self._determine_232_status(content_type, materials),
            'CustomerRef': item.get('project_number', '') or item.get('Project Number', ''),
            '_content_type': content_type,  # Hidden field for styling
        }

        return row

    def _calculate_quantities(self, qty_unit: str, quantity: float, weight: float, content_type: str) -> Tuple[str, str]:
        """
        Calculate Qty1 and Qty2 based on unit type.

        CBP Rule: ALL derivative rows get Qty2 = weight

        Logic from TariffMill:
        - Weight-only units (KG, G, T): Qty1 = weight, Qty2 = empty
        - Count-only units (NO, PCS, DOZ): Qty1 = quantity, Qty2 = empty
        - Dual units: Qty1 = quantity, Qty2 = weight
        - Derivative rows: Always include Qty2 = weight
        """
        weight_str = str(int(round(weight))) if weight > 0 else ''
        qty_str = str(int(round(quantity))) if quantity > 0 else ''

        # Weight-only units
        if qty_unit in ['KG', 'G', 'T', 'kg', 'g', 't']:
            return weight_str, ''

        # Count-only units
        elif qty_unit in ['NO', 'PCS', 'DOZ', 'no', 'pcs', 'doz']:
            # For derivative rows, CBP requires Qty2 = weight
            return qty_str, weight_str

        # Dual units
        else:
            return qty_str, weight_str

    def _determine_232_status(self, content_type: str, materials: Dict) -> str:
        """Determine 232 status for the row."""
        if content_type == 'non_232':
            return 'Non_232'
        elif content_type in ['steel', 'aluminum']:
            return f'232_{content_type.title()}'
        else:
            return ''

    def _export_to_excel(self, rows: List[Dict], output_path: Path):
        """Export rows to Excel with Section 232 formatting."""
        if not rows:
            return

        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Section 232 Export"

        # Write headers
        for col_idx, header in enumerate(self.EXPORT_COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        # Write data rows with formatting
        for row_idx, row_data in enumerate(rows, start=2):
            content_type = row_data.get('_content_type', '')

            # Determine font color based on material type
            color = self._get_material_color(content_type)
            font = Font(color=color)

            # Determine fill color for dual declarations
            fill = None
            if row_data.get('DualDeclaration'):
                fill = PatternFill(start_color='E1BEE7', end_color='E1BEE7', fill_type='solid')

            for col_idx, header in enumerate(self.EXPORT_COLUMNS, start=1):
                value = row_data.get(header, '')
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = font
                if fill:
                    cell.fill = fill

        # Auto-size columns
        for col_idx in range(1, len(self.EXPORT_COLUMNS) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 15

        # Set page setup for landscape
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.fitToWidth = 1

        # Save workbook
        wb.save(output_path)

    def _get_material_color(self, content_type: str) -> str:
        """Get RGB color code for material type."""
        color_map = {
            'steel': '4a4a4a',
            'aluminum': '6495ED',
            'copper': 'B87333',
            'wood': '8B4513',
            'auto': '2F4F4F',
            'non_232': 'FF0000',
        }
        return color_map.get(content_type, '000000')


def main():
    """Command-line interface for Section 232 exporter."""
    import sys
    from pathlib import Path

    if len(sys.argv) < 4:
        print("Usage: python section232_exporter.py <input_folder> <output_folder> <db_path>")
        sys.exit(1)

    input_folder = Path(sys.argv[1])
    output_folder = Path(sys.argv[2])
    db_path = Path(sys.argv[3])

    exporter = Section232Exporter(input_folder, output_folder, db_path)
    count = exporter.process_all()
    print(f"\nSection 232 export complete: {count} files processed")


if __name__ == "__main__":
    main()
